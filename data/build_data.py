"""
Generate two Excel workbooks (Hebrew RTL):
1) countries_sectors_items.xlsx  – Region / Country / Demographic-Sector / Gender / Daily-Color-Item
2) items_stores.xlsx             – Country / Item / Popular store(s) / Official website

Scope: Europe, North America (incl. all 50 US states), South America, Asia.
Sectors: condensed (2-3 demographic groups per country/state).
Genders: Female + Male.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------- Item catalog (gender-aware) ----------------------------------

FEMALE_ITEMS_GENERIC = [
    "צעיף",
    "סיכת שיער",
    "עגילים",
    "סיכת חזה (ברוש)",
    "צמיד",
    "תליון",
    "סרט שיער",
    "מטפחת ראש",
    "סיכת פרח לדש",
    "ציפורן (לק) - בקבוקון יומי",
]

MALE_ITEMS_GENERIC = [
    "עניבה",
    "מטפחת חזה",
    "סיכת דש (Lapel pin)",
    "חפתים",
    "פפיון",
    "כיפה",
    "צמיד יד",
    "רצועת שעון",
    "סרט כובע",
    "עט בכיס החולצה",
]

# Cultural item overrides per macro-culture
CULTURAL_ITEMS = {
    # macro-culture key: (female list, male list)
    "muslim": (
        ["חיג'אב", "סיכת חיג'אב", "צעיף ראש", "עגילים תואמים", "תליון", "צמיד"],
        ["כובע תקייה / כופי", "מטפחת חזה", "סיכת דש", "שעון יד עם רצועה צבעונית", "טבעת"],
    ),
    "jewish_religious": (
        ["מטפחת ראש (טישבנד)", "כובע (ברט)", "סיכת חזה צנועה", "צעיף", "תליון"],
        ["כיפה", "עניבה", "מטפחת חזה", "סיכת דש", "עט בכיס"],
    ),
    "hindu": (
        ["בינדי (נקודה במצח) צבעונית", "צמידי טבעת (צ'ורי)", "סארי-פין", "עגילים", "פרח לשיער"],
        ["מטפחת ראש (פגאדי)", "סיכת דש", "מחרוזת/קמע", "צעיף נחרו", "עט"],
    ),
    "sikh": (
        ["דופטה (צעיף ראש)", "צמידים", "עגילים", "סיכת שיער"],
        ["טורבן (דאסטאר) - עניבת ראש בצבע יומי", "קארה (צמיד פלדה) עם סרט צבעוני"],
    ),
    "buddhist_east_asia": (
        ["סיכת שיער", "צעיף משי", "סיכת קימונו (אובי-דומה)", "עגילים", "תליון"],
        ["עניבה", "מטפחת חזה", "סיכת דש", "רצועת שעון"],
    ),
    "indigenous_americas": (
        ["סרט שיער מסורתי", "עגילים", "מחרוזת חרוזים", "סיכת שיער"],
        ["סרט ראש מסורתי", "מחרוזת חרוזים", "סיכת דש"],
    ),
    "andean": (
        ["צעיף (לליקליה)", "כובע מסורתי", "סיכה (טופו)", "מחרוזת"],
        ["צ'ולו / כובע צמר", "צעיף", "סיכת דש"],
    ),
    "latin_general": (
        ["סיכת שיער", "צעיף", "עגילים", "פרח לשיער", "סיכת חזה"],
        ["עניבה", "מטפחת חזה", "סיכת דש", "כובע פנמה"],
    ),
    "african_diaspora": (
        ["מטפחת ראש (Gele/Headwrap)", "עגילים", "צמיד עץ/חרוזים", "סיכת שיער"],
        ["כובע קופיה", "צעיף", "סיכת דש"],
    ),
    "nordic_secular": (
        ["סיכת שיער", "צעיף", "עגילים", "תליון מינימליסטי"],
        ["עניבה", "מטפחת חזה", "סיכת דש", "עט"],
    ),
    "anglo_secular": (
        ["סיכת שיער", "צעיף", "עגילים", "תליון", "סיכת חזה"],
        ["עניבה", "מטפחת חזה", "סיכת דש", "חפתים", "עט"],
    ),
}

# ---------- Country -> sectors (condensed demographic) -------------------
# Each entry: country : [ (sector_name, culture_key), ... ]

EUROPE = {
    "אלבניה":            [("רוב אלבני (מוסלמי-חילוני)", "muslim"), ("מיעוט יווני אורתודוקסי", "anglo_secular")],
    "אנדורה":            [("רוב קטלאני", "anglo_secular"), ("מיעוט ספרדי/פורטוגלי", "latin_general")],
    "אוסטריה":           [("רוב אוסטרי-גרמני", "anglo_secular"), ("מיעוט טורקי/בלקני מוסלמי", "muslim")],
    "בלארוס":            [("רוב בלארוסי-אורתודוקסי", "anglo_secular"), ("מיעוט פולני-קתולי", "anglo_secular")],
    "בלגיה":             [("פלמים (דוברי הולנדית)", "nordic_secular"), ("ולונים (דוברי צרפתית)", "anglo_secular"), ("קהילה מוסלמית", "muslim")],
    "בוסניה והרצגובינה":  [("בושנאקים מוסלמים", "muslim"), ("סרבים אורתודוקסים", "anglo_secular"), ("קרואטים קתולים", "anglo_secular")],
    "בולגריה":           [("רוב בולגרי-אורתודוקסי", "anglo_secular"), ("מיעוט טורקי-מוסלמי", "muslim")],
    "קרואטיה":           [("רוב קרואטי-קתולי", "anglo_secular"), ("מיעוט סרבי-אורתודוקסי", "anglo_secular")],
    "קפריסין":            [("יוונים-קפריסאים", "anglo_secular"), ("טורקים-קפריסאים", "muslim")],
    "צ'כיה":              [("רוב צ'כי-חילוני", "nordic_secular"), ("מיעוט סלובקי", "anglo_secular")],
    "דנמרק":              [("רוב דני-לותרני", "nordic_secular"), ("קהילת מהגרים מוסלמים", "muslim")],
    "אסטוניה":            [("רוב אסטוני", "nordic_secular"), ("מיעוט רוסי-אורתודוקסי", "anglo_secular")],
    "פינלנד":             [("רוב פיני", "nordic_secular"), ("מיעוט שוודי-פיני", "nordic_secular")],
    "צרפת":               [("רוב צרפתי-חילוני", "anglo_secular"), ("קהילה מוסלמית (מגרב)", "muslim"), ("קהילה יהודית", "jewish_religious")],
    "גרמניה":             [("רוב גרמני", "anglo_secular"), ("מיעוט טורקי-מוסלמי", "muslim"), ("מהגרים מזרח-אירופים", "anglo_secular")],
    "יוון":               [("רוב יווני-אורתודוקסי", "anglo_secular"), ("מיעוט מוסלמי בתראקיה", "muslim")],
    "הונגריה":            [("רוב הונגרי", "anglo_secular"), ("מיעוט רומה (צוענים)", "anglo_secular")],
    "איסלנד":             [("רוב איסלנדי", "nordic_secular"), ("מהגרים מזרח-אירופים", "anglo_secular")],
    "אירלנד":             [("רוב אירי-קתולי", "anglo_secular"), ("קהילת מהגרים", "anglo_secular")],
    "איטליה":             [("רוב איטלקי-קתולי", "latin_general"), ("קהילת מהגרים צפון-אפריקאית", "muslim")],
    "קוסובו":             [("אלבנים מוסלמים", "muslim"), ("סרבים אורתודוקסים", "anglo_secular")],
    "לטביה":              [("רוב לטבי", "anglo_secular"), ("מיעוט רוסי", "anglo_secular")],
    "ליכטנשטיין":         [("רוב ליכטנשטייני-קתולי", "anglo_secular"), ("מהגרים גרמנים/אוסטרים", "anglo_secular")],
    "ליטא":               [("רוב ליטאי-קתולי", "anglo_secular"), ("מיעוט פולני/רוסי", "anglo_secular")],
    "לוקסמבורג":          [("רוב לוקסמבורגי", "anglo_secular"), ("מהגרים פורטוגלים", "latin_general")],
    "מלטה":               [("רוב מלטזי-קתולי", "anglo_secular"), ("מהגרים מצפון אפריקה", "muslim")],
    "מולדובה":            [("רוב מולדבי-אורתודוקסי", "anglo_secular"), ("מיעוט גגאוזי", "anglo_secular")],
    "מונקו":              [("רוב מונגאסקי/צרפתי", "anglo_secular"), ("קהילה איטלקית", "latin_general")],
    "מונטנגרו":           [("רוב מונטנגרי", "anglo_secular"), ("מיעוט בושנאקי-מוסלמי", "muslim")],
    "הולנד":              [("רוב הולנדי-חילוני", "nordic_secular"), ("קהילה מוסלמית (מרוקאית/טורקית)", "muslim"), ("מהגרים מסורינאם/אנטילים", "african_diaspora")],
    "צפון מקדוניה":       [("רוב מקדוני-אורתודוקסי", "anglo_secular"), ("מיעוט אלבני-מוסלמי", "muslim")],
    "נורבגיה":            [("רוב נורבגי", "nordic_secular"), ("עם הסאמי הילידי", "indigenous_americas")],
    "פולין":              [("רוב פולני-קתולי", "anglo_secular"), ("מהגרים אוקראינים", "anglo_secular")],
    "פורטוגל":            [("רוב פורטוגלי", "latin_general"), ("מהגרים מארצות אפריקה דוברות פורטוגלית", "african_diaspora")],
    "רומניה":             [("רוב רומני-אורתודוקסי", "anglo_secular"), ("מיעוט הונגרי/רומה", "anglo_secular")],
    "רוסיה":              [("רוב רוסי-אורתודוקסי", "anglo_secular"), ("מיעוטים מוסלמים (טטרים/קווקזים)", "muslim"), ("ילידי הצפון/סיביר", "indigenous_americas")],
    "סן מרינו":           [("רוב סן-מריני-קתולי", "latin_general"), ("מהגרים איטלקים", "latin_general")],
    "סרביה":              [("רוב סרבי-אורתודוקסי", "anglo_secular"), ("מיעוט בושנאקי/אלבני-מוסלמי", "muslim")],
    "סלובקיה":            [("רוב סלובקי-קתולי", "anglo_secular"), ("מיעוט הונגרי/רומה", "anglo_secular")],
    "סלובניה":            [("רוב סלובני-קתולי", "anglo_secular"), ("מהגרים מהבלקן", "anglo_secular")],
    "ספרד":               [("רוב ספרדי-קתולי", "latin_general"), ("קהילה קטלאנית/בסקית", "anglo_secular"), ("מהגרים מצפון אפריקה", "muslim")],
    "שבדיה":              [("רוב שבדי", "nordic_secular"), ("מהגרים מהמזה\"ת ומסומליה", "muslim")],
    "שווייץ":             [("דוברי גרמנית", "anglo_secular"), ("דוברי צרפתית", "anglo_secular"), ("דוברי איטלקית", "latin_general")],
    "אוקראינה":           [("רוב אוקראיני-אורתודוקסי", "anglo_secular"), ("מיעוט רוסי/טטרי", "anglo_secular")],
    "בריטניה":            [("רוב לבן-בריטי", "anglo_secular"), ("בריטים-אסיאתים (הודים/פקיסטנים)", "hindu"), ("בריטים-קריבים/אפריקאים", "african_diaspora")],
    "ותיקן":              [("הקהילה הכנסייתית-קתולית", "anglo_secular"), ("מבקרים/עובדים חילונים", "latin_general")],
}

SOUTH_AMERICA = {
    "ארגנטינה":           [("רוב ממוצא אירופי-קתולי", "latin_general"), ("ילידים (מאפוצ'ה, קצ'ואה)", "indigenous_americas")],
    "בוליביה":            [("רוב ילידי (קצ'ואה, איימרה)", "andean"), ("מסטיסוס", "latin_general")],
    "ברזיל":              [("ממוצא אירופי", "latin_general"), ("ממוצא אפרו-ברזילאי", "african_diaspora"), ("ילידים אמזוניים", "indigenous_americas")],
    "צ'ילה":              [("רוב מסטיסו", "latin_general"), ("עם המאפוצ'ה הילידי", "indigenous_americas")],
    "קולומביה":           [("רוב מסטיסו", "latin_general"), ("אפרו-קולומביאנים", "african_diaspora"), ("ילידי וייו/אמזון", "indigenous_americas")],
    "אקוודור":            [("רוב מסטיסו", "latin_general"), ("ילידי הקצ'ואה/שואר", "andean")],
    "גיאנה":              [("הודים-קריבים", "hindu"), ("אפרו-גיאנים", "african_diaspora")],
    "פרגוואי":            [("רוב מסטיסו דובר גוארני", "latin_general"), ("ילידי גוארני", "indigenous_americas")],
    "פרו":                [("רוב מסטיסו", "latin_general"), ("ילידי קצ'ואה ואיימרה", "andean")],
    "סורינאם":            [("הודים-סורינאמים", "hindu"), ("אפרו-סורינאמים (מארון/קריאולי)", "african_diaspora"), ("ג'אווה-סורינאמים", "muslim")],
    "אורוגוואי":          [("רוב ממוצא אירופי-חילוני", "anglo_secular"), ("אפרו-אורוגוואים", "african_diaspora")],
    "ונצואלה":            [("רוב מסטיסו", "latin_general"), ("אפרו-ונצואלים", "african_diaspora"), ("ילידי וייו", "indigenous_americas")],
}

NORTH_AMERICA_COUNTRIES = {
    "קנדה":               [("קנדים אנגלופונים", "anglo_secular"), ("קוויבקואים פרנקופונים", "anglo_secular"), ("עמים ילידיים (First Nations, Inuit, Métis)", "indigenous_americas")],
    "מקסיקו":             [("רוב מסטיסו", "latin_general"), ("ילידים (נאוואטל, מאיה)", "indigenous_americas")],
    "בליז":               [("מסטיסוס", "latin_general"), ("קריאולים אפרו-קריביים", "african_diaspora"), ("מאיה ילידים", "indigenous_americas")],
    "קוסטה ריקה":         [("רוב מסטיסו", "latin_general"), ("אפרו-קריביים בלימון", "african_diaspora")],
    "אל סלבדור":          [("רוב מסטיסו", "latin_general"), ("ילידי פיפיל", "indigenous_americas")],
    "גואטמלה":            [("ילידי מאיה", "indigenous_americas"), ("לאדינוס (מסטיסוס)", "latin_general")],
    "הונדורס":            [("רוב מסטיסו", "latin_general"), ("גריפונה (אפרו-קריבים)", "african_diaspora"), ("ילידי לנקה", "indigenous_americas")],
    "ניקרגואה":           [("רוב מסטיסו", "latin_general"), ("מיסקיטו ילידים/קריאולים", "indigenous_americas")],
    "פנמה":               [("רוב מסטיסו", "latin_general"), ("ילידי גונה/אמברה", "indigenous_americas"), ("אפרו-קריבים", "african_diaspora")],
    "אנטיגואה וברבודה":    [("אפרו-קריביים", "african_diaspora"), ("מהגרים בריטים/לבנים", "anglo_secular")],
    "בהאמה":              [("אפרו-בהאמים", "african_diaspora"), ("מיעוט לבן/הייטי", "latin_general")],
    "ברבדוס":             [("אפרו-ברבדוסים", "african_diaspora"), ("מיעוט הודי/לבן", "anglo_secular")],
    "קובה":               [("אפרו-קובנים", "african_diaspora"), ("ממוצא אירופי-ספרדי", "latin_general")],
    "דומיניקה":           [("אפרו-קריביים", "african_diaspora"), ("עם הקליננאגו הילידי", "indigenous_americas")],
    "הרפובליקה הדומיניקנית": [("מולאטים", "latin_general"), ("אפרו-דומיניקנים", "african_diaspora")],
    "גרנדה":              [("אפרו-קריבים", "african_diaspora"), ("מיעוט הודי/לבן", "anglo_secular")],
    "האיטי":              [("אפרו-האיטים", "african_diaspora"), ("מולאטים", "latin_general")],
    "ג'מייקה":            [("אפרו-ג'מייקנים", "african_diaspora"), ("הודים/סינים ג'מייקנים", "hindu")],
    "סנט קיטס ונוויס":     [("אפרו-קריבים", "african_diaspora"), ("מיעוט לבן/הודי", "anglo_secular")],
    "סנט לוסיה":          [("אפרו-קריבים", "african_diaspora"), ("מיעוט הודי/לבן", "anglo_secular")],
    "סנט וינסנט והגרנדינים": [("אפרו-קריבים", "african_diaspora"), ("קליננאגו ילידים", "indigenous_americas")],
    "טרינידד וטובגו":      [("הודים-טרינידדיים", "hindu"), ("אפרו-טרינידדיים", "african_diaspora")],
}

US_STATES = [
    "אלבמה", "אלסקה", "אריזונה", "ארקנסו", "קליפורניה", "קולורדו", "קונטיקט",
    "דלאוור", "פלורידה", "ג'ורג'יה", "הוואי", "איידהו", "אילינוי", "אינדיאנה",
    "איווה", "קנזס", "קנטקי", "לואיזיאנה", "מיין", "מרילנד", "מסצ'וסטס",
    "מישיגן", "מינסוטה", "מיסיסיפי", "מיזורי", "מונטנה", "נברסקה", "נבדה",
    "ניו המפשיר", "ניו ג'רזי", "ניו מקסיקו", "ניו יורק", "צפון קרוליינה",
    "צפון דקוטה", "אוהיו", "אוקלהומה", "אורגון", "פנסילבניה", "רוד איילנד",
    "דרום קרוליינה", "דרום דקוטה", "טנסי", "טקסס", "יוטה", "ורמונט",
    "וירג'יניה", "וושינגטון", "מערב וירג'יניה", "ויסקונסין", "ויומינג",
]

# US default sectors (varies slightly per state but using a common condensed scheme):
def us_state_sectors(state):
    # Customize for a few states with distinctive demographics
    special = {
        "הוואי":         [("אסיאתים-אמריקאים (יפנים/פיליפינים)", "buddhist_east_asia"), ("ילידי הוואי/האיים", "indigenous_americas"), ("לבנים", "anglo_secular")],
        "ניו מקסיקו":     [("היספאנים/לטינוס", "latin_general"), ("ילידים אינדיאנים (נבחו/פואבלו)", "indigenous_americas"), ("לבנים לא-היספאנים", "anglo_secular")],
        "אלסקה":         [("ילידי אלסקה (אינואיט/אתבסקאן)", "indigenous_americas"), ("לבנים", "anglo_secular")],
        "קליפורניה":      [("היספאנים/לטינוס", "latin_general"), ("אסיאתים-אמריקאים", "buddhist_east_asia"), ("לבנים לא-היספאנים", "anglo_secular")],
        "ניו יורק":       [("לבנים", "anglo_secular"), ("אפרו-אמריקאים", "african_diaspora"), ("היספאנים", "latin_general"), ("יהודים אורתודוקסים", "jewish_religious")],
        "טקסס":          [("היספאנים/לטינוס", "latin_general"), ("לבנים לא-היספאנים", "anglo_secular"), ("אפרו-אמריקאים", "african_diaspora")],
        "פלורידה":        [("היספאנים (קובנים/ונצואלים)", "latin_general"), ("אפרו-אמריקאים", "african_diaspora"), ("לבנים לא-היספאנים", "anglo_secular")],
        "מיסיסיפי":       [("אפרו-אמריקאים", "african_diaspora"), ("לבנים לא-היספאנים", "anglo_secular")],
        "מישיגן":         [("לבנים לא-היספאנים", "anglo_secular"), ("אפרו-אמריקאים", "african_diaspora"), ("אמריקאים-ערבים (דירבורן)", "muslim")],
        "מינסוטה":        [("לבנים סקנדינביים", "nordic_secular"), ("אפרו-אמריקאים (סומלים)", "muslim"), ("ילידים", "indigenous_americas")],
        "דרום דקוטה":     [("לבנים", "anglo_secular"), ("ילידים (לקוטה/דקוטה)", "indigenous_americas")],
        "צפון דקוטה":     [("לבנים", "anglo_secular"), ("ילידים", "indigenous_americas")],
        "אריזונה":        [("היספאנים/לטינוס", "latin_general"), ("ילידים (נבחו/הופי)", "indigenous_americas"), ("לבנים לא-היספאנים", "anglo_secular")],
        "אוקלהומה":       [("לבנים לא-היספאנים", "anglo_secular"), ("ילידים (צ'ירוקי/צ'וקטו)", "indigenous_americas")],
        "יוטה":          [("לבנים מורמונים (LDS)", "anglo_secular"), ("היספאנים/לטינוס", "latin_general")],
    }
    if state in special:
        return special[state]
    # default for most states
    return [
        ("לבנים לא-היספאנים", "anglo_secular"),
        ("אפרו-אמריקאים", "african_diaspora"),
        ("היספאנים/לטינוס", "latin_general"),
    ]

ASIA = {
    "אפגניסטן":           [("פשטונים מוסלמים", "muslim"), ("טאג'יקים/האזארים", "muslim")],
    "ארמניה":             [("רוב ארמני-אורתודוקסי", "anglo_secular"), ("מיעוט יזידי", "muslim")],
    "אזרבייג'ן":          [("רוב אזרי-מוסלמי", "muslim"), ("מיעוט רוסי/לזגי", "anglo_secular")],
    "בחריין":             [("ערבים סונים", "muslim"), ("ערבים שיעים", "muslim")],
    "בנגלדש":             [("רוב בנגלי-מוסלמי", "muslim"), ("מיעוט הינדי", "hindu")],
    "בהוטן":              [("דרוקפה בודהיסטים", "buddhist_east_asia"), ("לוטשמפה הינדים (ממוצא נפאלי)", "hindu")],
    "ברוניי":             [("מאלאים מוסלמים", "muslim"), ("סינים-בודהיסטים", "buddhist_east_asia")],
    "קמבודיה":            [("חמרים בודהיסטים", "buddhist_east_asia"), ("צ'אם מוסלמים", "muslim")],
    "סין":                [("האן (רוב)", "buddhist_east_asia"), ("מיעוטים מוסלמים (הוּיי/אויגורים)", "muslim"), ("טיבטים בודהיסטים", "buddhist_east_asia")],
    "גאורגיה":            [("רוב גאורגי-אורתודוקסי", "anglo_secular"), ("מיעוט אזרי/ארמני", "muslim")],
    "הודו":               [("הינדים", "hindu"), ("מוסלמים", "muslim"), ("סיקים", "sikh")],
    "אינדונזיה":          [("רוב ג'אווני-מוסלמי", "muslim"), ("מיעוט בלינזי-הינדי", "hindu"), ("סינים-אינדונזים", "buddhist_east_asia")],
    "איראן":              [("פרסים שיעים", "muslim"), ("מיעוטים (אזרים/כורדים/ערבים)", "muslim")],
    "עיראק":              [("ערבים שיעים", "muslim"), ("ערבים סונים", "muslim"), ("כורדים", "muslim")],
    "ישראל":              [("יהודים חילונים", "jewish_religious"), ("יהודים דתיים/חרדים", "jewish_religious"), ("ערבים-מוסלמים", "muslim"), ("ערבים-נוצרים/דרוזים", "muslim")],
    "יפן":                [("יפנים (שינטו-בודהיסטים)", "buddhist_east_asia"), ("מיעוט אינו ילידי", "indigenous_americas")],
    "ירדן":               [("ערבים ירדנים", "muslim"), ("פלסטינים-ירדנים", "muslim")],
    "קזחסטן":             [("קזחים מוסלמים", "muslim"), ("מיעוט רוסי-אורתודוקסי", "anglo_secular")],
    "כווית":              [("ערבים סונים", "muslim"), ("ערבים שיעים", "muslim")],
    "קירגיזסטן":          [("קירגיזים מוסלמים", "muslim"), ("מיעוט רוסי", "anglo_secular")],
    "לאוס":               [("לאו בודהיסטים", "buddhist_east_asia"), ("המונג/שבטי ההר", "indigenous_americas")],
    "לבנון":              [("מוסלמים סונים", "muslim"), ("מוסלמים שיעים", "muslim"), ("נוצרים מארונים", "anglo_secular"), ("דרוזים", "muslim")],
    "מלזיה":              [("מאלאים מוסלמים", "muslim"), ("סינים-מלזים", "buddhist_east_asia"), ("הודים-מלזים", "hindu")],
    "מלדיביים":           [("רוב מלדיבי-מוסלמי", "muslim"), ("מהגרי עבודה (בנגלים/הודים)", "muslim")],
    "מונגוליה":           [("חלחים בודהיסטים", "buddhist_east_asia"), ("קזחים מוסלמים במערב", "muslim")],
    "מיאנמר":             [("בורמזים בודהיסטים", "buddhist_east_asia"), ("רוהינגיה מוסלמים", "muslim"), ("שאן/קארן", "indigenous_americas")],
    "נפאל":               [("הינדים (פהאדי)", "hindu"), ("בודהיסטים (נווארי/שרפה)", "buddhist_east_asia")],
    "צפון קוריאה":        [("קוריאנים (רוב)", "buddhist_east_asia"), ("מיעוט סיני קטן", "buddhist_east_asia")],
    "עומאן":              [("ערבים אבאדים", "muslim"), ("ערבים סונים/שיעים", "muslim")],
    "פקיסטן":             [("פנג'אבים סונים", "muslim"), ("פשטונים/בלוצ'ים", "muslim"), ("מיעוט הינדי/מסיחי", "hindu")],
    "פלסטין":             [("מוסלמים סונים", "muslim"), ("נוצרים פלסטינים", "anglo_secular")],
    "הפיליפינים":         [("רוב קתולי", "latin_general"), ("מוסלמי-מורו במינדנאו", "muslim"), ("ילידי קורדילרה", "indigenous_americas")],
    "קטר":                [("ערבים סונים מקומיים", "muslim"), ("מהגרי עבודה (הודים/בנגלים/פיליפינים)", "hindu")],
    "ערב הסעודית":        [("ערבים סונים", "muslim"), ("ערבים שיעים (מזרח)", "muslim")],
    "סינגפור":            [("סינים-סינגפורים", "buddhist_east_asia"), ("מאלאים מוסלמים", "muslim"), ("הודים-סינגפורים", "hindu")],
    "דרום קוריאה":        [("קוריאנים (נוצרים/בודהיסטים)", "buddhist_east_asia"), ("מהגרי עבודה אסיאתים", "buddhist_east_asia")],
    "סרי לנקה":           [("סינהלים בודהיסטים", "buddhist_east_asia"), ("טמילים הינדים", "hindu"), ("מוסלמים", "muslim")],
    "סוריה":              [("ערבים סונים", "muslim"), ("עלווים/דרוזים/נוצרים", "muslim"), ("כורדים", "muslim")],
    "טייוואן":            [("הוקלו/האקה (רוב סיני)", "buddhist_east_asia"), ("עמים אוסטרונזים ילידיים", "indigenous_americas")],
    "טג'יקיסטן":          [("טאג'יקים מוסלמים", "muslim"), ("מיעוט אוזבקי/רוסי", "muslim")],
    "תאילנד":             [("תאים בודהיסטים", "buddhist_east_asia"), ("מאלאים-מוסלמים בדרום", "muslim"), ("שבטי הר ילידיים", "indigenous_americas")],
    "טימור-לסטה":         [("טטום קתולים", "latin_general"), ("ילידי המאלאי-פולינזי", "indigenous_americas")],
    "טורקיה":             [("טורקים סונים", "muslim"), ("כורדים", "muslim"), ("עלווים", "muslim")],
    "טורקמניסטן":         [("טורקמנים מוסלמים", "muslim"), ("מיעוט אוזבקי/רוסי", "muslim")],
    "איחוד האמירויות":     [("אמירתים סונים", "muslim"), ("מהגרי עבודה (הודים/פקיסטנים/פיליפינים)", "hindu")],
    "אוזבקיסטן":          [("אוזבקים מוסלמים", "muslim"), ("מיעוט קרקלפק/רוסי", "muslim")],
    "וייטנאם":            [("ויאט (קין) בודהיסטים", "buddhist_east_asia"), ("שבטי הר ילידיים", "indigenous_americas"), ("חמרים-ויאטנמים", "buddhist_east_asia")],
    "תימן":               [("ערבים סונים", "muslim"), ("ערבים זיידים-שיעים", "muslim")],
}

# ---------- Item assignment ---------------------------------------------

def items_for(culture_key):
    """Return (female_list, male_list) for the culture."""
    if culture_key in CULTURAL_ITEMS:
        f, m = CULTURAL_ITEMS[culture_key]
        return f, m
    return FEMALE_ITEMS_GENERIC, MALE_ITEMS_GENERIC

def pick_item(items, idx):
    return items[idx % len(items)]

# ---------- Store catalog -----------------------------------------------
# country -> [ (chain_name, official_website) ... ]
STORES = {
    # Europe
    "אלבניה":            [("TEG (Tirana East Gate)", "https://teg.al/"), ("QTU (Tirana shopping center)", "https://www.qtu.al/")],
    "אנדורה":            [("Pyrénées Andorra", "https://www.pyrenees.ad/"), ("Andorra2000", "https://www.andorra2000.com/")],
    "אוסטריה":           [("Peek & Cloppenburg AT", "https://www.peek-cloppenburg.at/"), ("Steffl Vienna", "https://www.steffl-vienna.at/")],
    "בלארוס":            [("GUM Minsk", "https://gum.by/"), ("Galleria Minsk", "https://galleria.by/")],
    "בלגיה":             [("Inno", "https://www.inno.be/"), ("Galeria Inno", "https://www.galeria-inno.be/"), ("Hema BE", "https://www.hema.be/")],
    "בוסניה והרצגובינה":  [("Sarajevo City Center", "https://www.scc.ba/"), ("Bingo BH", "https://bingotuzla.ba/")],
    "בולגריה":           [("CCS (Central Cooperative Store)", "https://www.tsum.bg/"), ("Paradise Center Sofia", "https://paradise-center.com/")],
    "קרואטיה":           [("NaMa", "https://www.nama.hr/"), ("Müller HR", "https://www.mueller.hr/")],
    "קפריסין":            [("Debenhams Cyprus", "https://www.debenhams.com.cy/"), ("Public Cyprus", "https://www.public.com.cy/")],
    "צ'כיה":              [("Kotva Department Store", "https://www.od-kotva.cz/"), ("Palladium Prague", "https://www.palladiumpraha.cz/")],
    "דנמרק":              [("Magasin du Nord", "https://www.magasin.dk/"), ("Illum", "https://www.illum.dk/")],
    "אסטוניה":            [("Kaubamaja", "https://www.kaubamaja.ee/"), ("Stockmann Tallinn", "https://stockmann.ee/")],
    "פינלנד":             [("Stockmann", "https://www.stockmann.com/"), ("Sokos", "https://www.sokos.fi/")],
    "צרפת":               [("Galeries Lafayette", "https://www.galerieslafayette.com/"), ("Le Printemps", "https://www.printemps.com/"), ("BHV Marais", "https://www.bhv.fr/")],
    "גרמניה":             [("KaDeWe", "https://www.kadewe.de/"), ("Galeria Karstadt Kaufhof", "https://www.galeria.de/"), ("Breuninger", "https://www.breuninger.com/")],
    "יוון":               [("Attica Department Stores", "https://www.atticadps.gr/"), ("Notos Home", "https://www.notos.gr/")],
    "הונגריה":            [("Westend Budapest", "https://www.westend.hu/"), ("Mammut", "https://www.mammut.hu/")],
    "איסלנד":             [("Hagkaup", "https://www.hagkaup.is/"), ("Kringlan", "https://www.kringlan.is/")],
    "אירלנד":             [("Brown Thomas", "https://www.brownthomas.com/"), ("Arnotts", "https://www.arnotts.ie/")],
    "איטליה":             [("La Rinascente", "https://www.rinascente.it/"), ("Coin", "https://www.coin.it/"), ("OVS", "https://www.ovs.it/")],
    "קוסובו":             [("Albi Mall Pristina", "https://albimall.com/"), ("Prishtina Mall", "https://prishtinamall.com/")],
    "לטביה":              [("Stockmann Riga", "https://www.stockmann.lv/"), ("Galerija Centrs", "https://www.galerijacentrs.lv/")],
    "ליכטנשטיין":         [("Huber Fine Watches & Jewellery", "https://www.huber.li/"), ("Manor Liechtenstein (via Switzerland)", "https://www.manor.ch/")],
    "ליטא":               [("Akropolis", "https://www.akropolis.lt/"), ("Ozas Vilnius", "https://www.ozas.lt/")],
    "לוקסמבורג":          [("Galeries Lafayette Luxembourg", "https://www.galerieslafayette.lu/"), ("Auchan Luxembourg", "https://www.auchan.lu/")],
    "מלטה":               [("The Point Sliema", "https://www.thepointmalta.com/"), ("Debenhams Malta", "https://www.debenhams.com.mt/")],
    "מולדובה":            [("UNIC", "https://unic.md/"), ("Shopping MallDova", "https://malldova.md/")],
    "מונקו":              [("Métropole Shopping Monte-Carlo", "https://www.metropoleshoppingmontecarlo.com/"), ("One Monte-Carlo", "https://www.onemontecarlo.com/")],
    "מונטנגרו":           [("Delta City Podgorica", "https://deltacity.me/"), ("Mall of Montenegro", "https://mallofmontenegro.com/")],
    "הולנד":              [("De Bijenkorf", "https://www.debijenkorf.nl/"), ("HEMA", "https://www.hema.nl/"), ("V&D successor: Etos/Wehkamp", "https://www.wehkamp.nl/")],
    "צפון מקדוניה":       [("Skopje City Mall", "https://skopjecitymall.mk/"), ("Ramstore Macedonia", "https://www.ramstore.mk/")],
    "נורבגיה":            [("Steen & Strøm", "https://www.steenogstrom.no/"), ("Glasmagasinet", "https://www.glasmagasinet.no/")],
    "פולין":              [("CH Arkadia", "https://www.arkadia.com.pl/"), ("Galeria Mokotów", "https://galeriamokotow.pl/"), ("Vitkac", "https://www.vitkac.com/")],
    "פורטוגל":            [("El Corte Inglés Lisboa", "https://www.elcorteingles.pt/"), ("Continente Modelo", "https://www.continente.pt/")],
    "רומניה":             [("Cocor Department Store", "https://www.cocor.ro/"), ("AFI Cotroceni", "https://www.afibucuresti.ro/")],
    "רוסיה":              [("GUM Moscow", "https://gum.ru/"), ("TSUM", "https://www.tsum.ru/"), ("Wildberries", "https://www.wildberries.ru/")],
    "סן מרינו":           [("Atlante Shopping San Marino", "https://www.atlantesanmarino.com/"), ("Azzurro Shopping", "https://www.azzurroshopping.com/")],
    "סרביה":              [("Ušće Shopping Center", "https://www.usceshoppingcenter.rs/"), ("Rajićeva Belgrade", "https://www.rajiceva.rs/")],
    "סלובקיה":            [("Aupark Bratislava", "https://www.aupark-bratislava.sk/"), ("Eurovea", "https://www.eurovea.com/")],
    "סלובניה":            [("Maxi Ljubljana (Nama)", "https://www.nama.si/"), ("BTC City", "https://www.btc-city.com/")],
    "ספרד":               [("El Corte Inglés", "https://www.elcorteingles.es/"), ("Mango", "https://shop.mango.com/"), ("Zara (Inditex)", "https://www.zara.com/")],
    "שבדיה":              [("Åhléns", "https://www.ahlens.se/"), ("NK (Nordiska Kompaniet)", "https://www.nk.se/")],
    "שווייץ":             [("Manor", "https://www.manor.ch/"), ("Globus", "https://www.globus.ch/"), ("Jelmoli", "https://www.jelmoli.ch/")],
    "אוקראינה":           [("TSUM Kyiv", "https://tsum.ua/"), ("Ocean Plaza", "https://oceanplaza.com.ua/")],
    "בריטניה":            [("John Lewis", "https://www.johnlewis.com/"), ("Marks & Spencer", "https://www.marksandspencer.com/"), ("Harrods", "https://www.harrods.com/"), ("Selfridges", "https://www.selfridges.com/")],
    "ותיקן":              [("Vatican Library Boutique", "https://www.museivaticani.va/"), ("Libreria Editrice Vaticana", "https://www.libreriaeditricevaticana.va/")],

    # South America
    "ארגנטינה":           [("Falabella Argentina", "https://www.falabella.com.ar/"), ("Galerías Pacífico", "https://galeriaspacifico.com.ar/"), ("Mercado Libre", "https://www.mercadolibre.com.ar/")],
    "בוליביה":            [("Hipermaxi", "https://hipermaxi.com/"), ("Casa Ideal La Paz", "https://casaideal.com.bo/")],
    "ברזיל":              [("Lojas Renner", "https://www.lojasrenner.com.br/"), ("Riachuelo", "https://www.riachuelo.com.br/"), ("Magazine Luiza", "https://www.magazineluiza.com.br/"), ("Americanas", "https://www.americanas.com.br/")],
    "צ'ילה":              [("Falabella", "https://www.falabella.com/"), ("Paris Cencosud", "https://www.paris.cl/"), ("Ripley", "https://simple.ripley.cl/")],
    "קולומביה":           [("Éxito", "https://www.exito.com/"), ("Falabella Colombia", "https://www.falabella.com.co/")],
    "אקוודור":            [("De Prati", "https://www.deprati.com.ec/"), ("Etafashion", "https://www.etafashion.com/")],
    "גיאנה":              [("Giftland Mall", "https://giftlandmall.com/"), ("Massy Stores Guyana", "https://www.massystoresgy.com/")],
    "פרגוואי":            [("Shopping del Sol", "https://shoppingdelsol.com.py/"), ("Casa Rica", "https://www.casarica.com.py/")],
    "פרו":                [("Saga Falabella", "https://www.falabella.com.pe/"), ("Ripley Perú", "https://simple.ripley.com.pe/"), ("Oechsle", "https://www.oechsle.pe/")],
    "סורינאם":            [("Hermitage Mall Paramaribo", "https://hermitagemall.com/"), ("Choi's Supermarket", "https://choissupermarket.com/")],
    "אורוגוואי":          [("Tienda Inglesa", "https://www.tiendainglesa.com.uy/"), ("Montevideo Shopping", "https://www.montevideoshopping.com.uy/")],
    "ונצואלה":            [("Traki", "https://www.tiendastraki.com.ve/"), ("Beco", "https://beco.com.ve/")],

    # North America (countries)
    "קנדה":               [("Hudson's Bay", "https://www.thebay.com/"), ("Holt Renfrew", "https://www.holtrenfrew.com/"), ("Simons", "https://www.simons.ca/"), ("Indigo", "https://www.indigo.ca/")],
    "מקסיקו":             [("Liverpool", "https://www.liverpool.com.mx/"), ("Palacio de Hierro", "https://www.elpalaciodehierro.com/"), ("Sears México", "https://www.sears.com.mx/"), ("Coppel", "https://www.coppel.com/")],
    "בליז":               [("Brodie's Department Store", "https://www.brodiesbelize.com/"), ("Mirab Belize", "https://mirab.bz/")],
    "קוסטה ריקה":         [("Universal", "https://www.universalcr.com/"), ("Multiplaza Escazú", "https://www.multiplaza.com/")],
    "אל סלבדור":          [("Simán", "https://www.siman.com/"), ("Multi Plaza", "https://multiplazasv.com/")],
    "גואטמלה":            [("Siman Guatemala", "https://www.siman.com/"), ("Cemaco", "https://www.cemaco.com/")],
    "הונדורס":            [("Diunsa", "https://www.diunsa.hn/"), ("Carrion", "https://www.carrion.hn/")],
    "ניקרגואה":           [("Siman Nicaragua", "https://www.siman.com/"), ("La Colonia", "https://lacolonia.com/")],
    "פנמה":               [("Felix B. Maduro", "https://www.felixmaduro.com/"), ("Multiplaza Pacific", "https://multiplaza.com/panama/")],
    "אנטיגואה וברבודה":    [("Epicurean Fine Foods & Pharmacy", "https://www.epicureanantigua.com/"), ("Townhouse Megastore", "https://townhouse.ag/")],
    "בהאמה":              [("John Bull", "https://www.johnbull.com/"), ("Solomon's Lucaya", "https://abacomarkets.com/")],
    "ברבדוס":             [("Cave Shepherd", "https://caveshepherd.com/"), ("Massy Stores Barbados", "https://www.massystoresbb.com/")],
    "קובה":               [("Tiendas Caribe", "https://www.tiendascaribe.cu/"), ("La Época Havana", "https://www.lcweb.cu/")],
    "דומיניקה":           [("Whitchurch IGA", "https://whitchurch.com/"), ("Astaphan's", "https://astaphans.com/")],
    "הרפובליקה הדומיניקנית": [("Sirena", "https://www.sirena.do/"), ("Jumbo", "https://www.jumbo.com.do/"), ("Plaza Lama", "https://plazalama.com.do/")],
    "גרנדה":              [("Spiceland Mall", "https://spicelandmall.com/"), ("Excel Plaza", "https://excelplaza.com/")],
    "האיטי":              [("Caribbean Market Haiti", "https://www.caribmarket.com/"), ("Eagle Market", "https://eaglemarket.ht/")],
    "ג'מייקה":            [("Sovereign Centre", "https://sovereigncentre.com/"), ("Lee's Fifth Avenue", "https://lees5thave.com/")],
    "סנט קיטס ונוויס":     [("RAMS Supermarkets", "https://ramsstkitts.com/"), ("Best Buy Furniture & Variety", "https://bestbuysknv.com/")],
    "סנט לוסיה":          [("Massy Stores St Lucia", "https://www.massystoreslc.com/"), ("Bay Walk Mall", "https://baywalkstlucia.com/")],
    "סנט וינסנט והגרנדינים": [("Massy Stores SVG", "https://www.massystoresvc.com/"), ("Greaves Department Store", "https://greavesltd.com/")],
    "טרינידד וטובגו":      [("Pricesmart Trinidad", "https://www.pricesmart.com/"), ("Massy Stores Trinidad", "https://massystorestt.com/")],

    # Asia
    "אפגניסטן":           [("Roshan Plaza Kabul", "https://www.roshan.af/"), ("Gulbahar Center", "https://gulbaharcenter.af/")],
    "ארמניה":             [("Yerevan Mall", "https://yerevanmall.am/"), ("Dalma Garden Mall", "https://www.dalmagardenmall.am/")],
    "אזרבייג'ן":          [("Park Bulvar", "https://parkbulvar.az/"), ("28 Mall", "https://www.28mall.az/")],
    "בחריין":             [("Bahrain City Centre", "https://www.citycentrebahrain.com/"), ("The Avenues Bahrain", "https://theavenues.bh/")],
    "בנגלדש":             [("Aarong", "https://www.aarong.com/"), ("Bashundhara City", "https://www.bashundhara-city.com/")],
    "בהוטן":              [("Tashi Commercial Corporation", "https://tashigroup.bt/"), ("8-Eleven Thimphu (Bhutan)", "https://www.8-eleven.com.bt/")],
    "ברוניי":             [("The Mall Gadong", "https://themallbrunei.com/"), ("Hua Ho Department Store", "https://www.huaho.com.bn/")],
    "קמבודיה":            [("Aeon Mall Phnom Penh", "https://www.aeonmallphnompenh.com/"), ("Lucky Department Store", "https://luckysupermarket.com.kh/")],
    "סין":                [("Taobao / Tmall", "https://www.taobao.com/"), ("JD.com", "https://www.jd.com/"), ("Wangfujing Department Store", "https://www.wangfujing.com/")],
    "גאורגיה":            [("Tbilisi Mall", "https://tbilisimall.com/"), ("Galleria Tbilisi", "https://galleriatbilisi.ge/")],
    "הודו":               [("Tanishq (Tata)", "https://www.tanishq.co.in/"), ("FabIndia", "https://www.fabindia.com/"), ("Reliance Trends", "https://www.reliancetrends.com/"), ("Myntra", "https://www.myntra.com/")],
    "אינדונזיה":          [("Sogo Indonesia", "https://www.sogo.co.id/"), ("Matahari Department Store", "https://www.matahari.com/"), ("Tokopedia", "https://www.tokopedia.com/")],
    "איראן":              [("Hyperstar Iran", "https://www.hyperstariran.com/"), ("Palladium Mall Tehran", "https://palladiummall.com/")],
    "עיראק":              [("Mansour Mall Baghdad", "https://mansourmall.com/"), ("Majidi Mall Erbil", "https://majidi-mall.com/")],
    "ישראל":              [("Castro", "https://www.castro.com/"), ("FOX", "https://www.fox.co.il/"), ("Renuar", "https://www.renuar.co.il/"), ("Padani (תכשיטים)", "https://www.padani.co.il/"), ("Hamashbir 365", "https://www.hamashbir365.co.il/")],
    "יפן":                [("Isetan Mitsukoshi", "https://www.imhds.co.jp/"), ("Takashimaya", "https://www.takashimaya.co.jp/"), ("Daimaru Matsuzakaya", "https://www.daimaru.co.jp/"), ("Muji", "https://www.muji.com/jp/")],
    "ירדן":               [("Safeway Jordan", "https://www.safeway.jo/"), ("Abdali Mall", "https://www.abdali-mall.com/")],
    "קזחסטן":             [("Mega Center Almaty", "https://mega.kz/"), ("Esentai Mall", "https://www.esentaimall.com/")],
    "כווית":              [("The Avenues Kuwait", "https://www.the-avenues.com/"), ("360 Mall", "https://www.360mall.com/")],
    "קירגיזסטן":          [("Tsum Bishkek", "https://tsum.kg/"), ("Bishkek Park", "https://www.bishkekpark.kg/")],
    "לאוס":               [("Vientiane Center", "https://vientianecenter.la/"), ("Talat Sao Mall", "https://talatsaomall.com/")],
    "לבנון":              [("ABC Department Stores", "https://www.abc.com.lb/"), ("Aïshti", "https://www.aishti.com/")],
    "מלזיה":              [("Parkson", "https://www.parkson.com.my/"), ("Isetan KL", "https://www.isetankl.com.my/"), ("Lazada Malaysia", "https://www.lazada.com.my/")],
    "מלדיביים":           [("STO People's Choice", "https://www.sto.mv/"), ("Big Mart Maldives", "https://bigmart.mv/")],
    "מונגוליה":           [("State Department Store Ulaanbaatar", "https://www.sdepartmentstore.com/"), ("Shangri-La Mall UB", "https://www.shangri-la.com/ulaanbaatar/")],
    "מיאנמר":             [("Junction City Yangon", "https://www.junctioncity.com.mm/"), ("City Mart", "https://citymart.com.mm/")],
    "נפאל":               [("Bhat-Bhateni Supermarket", "https://www.bbsm.com.np/"), ("Civil Mall Kathmandu", "https://civilmall.com.np/")],
    "צפון קוריאה":        [("Pyongyang Department Store No.1 (state-run)", "https://kcna.kp/")],
    "עומאן":              [("Lulu Hypermarket Oman", "https://www.luluhypermarket.com/"), ("Muscat Grand Mall", "https://www.muscatgrandmall.com/")],
    "פקיסטן":             [("Khaadi", "https://pk.khaadi.com/"), ("Junaid Jamshed (J.)", "https://www.junaidjamshed.com/"), ("Sapphire", "https://pk.sapphireonline.pk/")],
    "פלסטין":             [("Bravo Supermarket", "https://www.bravo.ps/"), ("Plaza Mall Ramallah", "https://plazamall.ps/")],
    "הפיליפינים":         [("SM Department Store", "https://www.smretail.com/"), ("Robinsons Department Store", "https://www.robinsonsdepartmentstore.com.ph/"), ("Rustan's", "https://www.rustans.com/")],
    "קטר":                [("Doha Festival City", "https://www.dohafestivalcity.com/"), ("Villaggio Mall", "https://www.villaggioqatar.com/"), ("Lulu Hypermarket Qatar", "https://www.luluhypermarket.com/")],
    "ערב הסעודית":        [("Centrepoint", "https://www.centrepointstores.com/"), ("Al-Othaim Markets", "https://othaimmarkets.com/"), ("Mall of Arabia Jeddah", "https://www.mallofarabia.com.sa/")],
    "סינגפור":            [("Takashimaya Singapore", "https://www.takashimaya.com.sg/"), ("Tangs", "https://www.tangs.com/"), ("Mustafa Centre", "https://www.mustafa.com.sg/")],
    "דרום קוריאה":        [("Lotte Department Store", "https://www.lotteshopping.com/"), ("Shinsegae", "https://www.shinsegae.com/"), ("Hyundai Department Store", "https://www.thehyundai.com/")],
    "סרי לנקה":           [("Odel", "https://odel.lk/"), ("Cool Planet", "https://www.coolplanet.lk/"), ("House of Fashion", "https://www.houseoffashionsl.com/")],
    "סוריה":              [("Shahba Mall Aleppo", "https://shahbamall.com/"), ("Cham City Center", "https://chamcitycenter.com/")],
    "טייוואן":            [("SOGO Taipei", "https://www.sogo.com.tw/"), ("Shin Kong Mitsukoshi", "https://www.skm.com.tw/"), ("PChome", "https://www.pchome.com.tw/")],
    "טג'יקיסטן":          [("Dushanbe Mall", "https://dushanbemall.tj/"), ("Auchan / Korzinka Dushanbe", "https://korzinka.tj/")],
    "תאילנד":             [("Central Department Store", "https://www.central.co.th/"), ("The Mall Group", "https://www.themallgroup.com/"), ("ICONSIAM", "https://www.iconsiam.com/")],
    "טימור-לסטה":         [("Timor Plaza", "https://timorplaza.com/"), ("Kmanek Supermarket", "https://kmanek.tl/")],
    "טורקיה":             [("Boyner", "https://www.boyner.com.tr/"), ("Beymen", "https://www.beymen.com/"), ("LC Waikiki", "https://www.lcwaikiki.com/")],
    "טורקמניסטן":         [("Berkarar Shopping Center", "https://berkarar.tm/"), ("Ashgabat Central Market", "https://ashgabatmarket.tm/")],
    "איחוד האמירויות":     [("The Dubai Mall", "https://thedubaimall.com/"), ("Mall of the Emirates", "https://www.malloftheemirates.com/"), ("Lulu Hypermarket UAE", "https://www.luluhypermarket.com/"), ("Centrepoint UAE", "https://www.centrepointstores.com/")],
    "אוזבקיסטן":          [("Samarqand Darvoza Mall", "https://samarqanddarvoza.uz/"), ("Compass Mall Tashkent", "https://compassmall.uz/")],
    "וייטנאם":            [("Vincom Center", "https://vincom.com.vn/"), ("Lotte Department Store Hanoi", "https://lottedepartmentstore.com.vn/"), ("AEON Mall Vietnam", "https://aeonmall-vietnam.com/")],
    "תימן":               [("Hayel Saeed Anam Group", "https://hsagroup.com/"), ("Sana'a City Center", "https://sanacitycenter.com/")],
}

# Default US chains
US_DEFAULT_STORES = [
    ("Target", "https://www.target.com/"),
    ("Walmart", "https://www.walmart.com/"),
    ("Macy's", "https://www.macys.com/"),
    ("Nordstrom", "https://www.nordstrom.com/"),
    ("Kohl's", "https://www.kohls.com/"),
    ("JCPenney", "https://www.jcpenney.com/"),
    ("Amazon", "https://www.amazon.com/"),
    ("Etsy (handmade)", "https://www.etsy.com/"),
    ("Tiffany & Co. (jewelry)", "https://www.tiffany.com/"),
]

US_STATE_LOCAL = {
    "ניו יורק":   [("Macy's Herald Square", "https://www.macys.com/"), ("Saks Fifth Avenue", "https://www.saksfifthavenue.com/"), ("Bloomingdale's", "https://www.bloomingdales.com/")],
    "קליפורניה":  [("Nordstrom CA", "https://www.nordstrom.com/"), ("Macy's", "https://www.macys.com/"), ("Ross Dress for Less", "https://www.rossstores.com/")],
    "טקסס":      [("H-E-B (general)", "https://www.heb.com/"), ("Dillard's", "https://www.dillards.com/"), ("Neiman Marcus (Dallas HQ)", "https://www.neimanmarcus.com/")],
    "פלורידה":    [("Dillard's", "https://www.dillards.com/"), ("Macy's Florida", "https://www.macys.com/"), ("Belk", "https://www.belk.com/")],
    "הוואי":     [("Ala Moana Center", "https://www.alamoanacenter.com/"), ("Macy's Hawaii", "https://www.macys.com/")],
    "אלסקה":     [("Nordstrom Anchorage", "https://www.nordstrom.com/"), ("Fred Meyer", "https://www.fredmeyer.com/")],
}

# ---------- Helpers ----------------------------------------------------

def stores_for(country):
    if country in STORES:
        return STORES[country]
    return []

def us_stores_for(state):
    base = US_STATE_LOCAL.get(state, [])
    return base + US_DEFAULT_STORES

# ---------- Build main sheet -------------------------------------------

def build_main_workbook(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "מדינות-מגזרים-פריטים"
    ws.sheet_view.rightToLeft = True

    headers = ["אזור", "מדינה / מדינת ארה\"ב", "מגזר דמוגרפי", "מגדר", "פריט יומי להחלפה לפי צבע", "הערה"]
    ws.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="305496")
    align = Alignment(horizontal="right", vertical="center", wrap_text=True)
    border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    for col_idx, _ in enumerate(headers, 1):
        c = ws.cell(row=1, column=col_idx)
        c.font = header_font
        c.fill = header_fill
        c.alignment = align
        c.border = border

    def write_block(region, mapping):
        for country, sectors in mapping.items():
            for sector_idx, (sector_name, culture) in enumerate(sectors):
                f_items, m_items = items_for(culture)
                female_item = pick_item(f_items, sector_idx)
                male_item = pick_item(m_items, sector_idx)
                ws.append([region, country, sector_name, "אישה",  female_item, "החלפה יומית לפי צבע - תואם לבוש/אירוע"])
                ws.append([region, country, sector_name, "גבר",   male_item,   "החלפה יומית לפי צבע - תואם לבוש/אירוע"])

    write_block("אירופה", EUROPE)
    write_block("אמריקה הצפונית", NORTH_AMERICA_COUNTRIES)
    # USA states as separate "region" rows
    for state in US_STATES:
        sectors = us_state_sectors(state)
        for sector_idx, (sector_name, culture) in enumerate(sectors):
            f_items, m_items = items_for(culture)
            ws.append(["אמריקה הצפונית - ארה\"ב", f"ארה\"ב - {state}", sector_name, "אישה", pick_item(f_items, sector_idx), "החלפה יומית לפי צבע"])
            ws.append(["אמריקה הצפונית - ארה\"ב", f"ארה\"ב - {state}", sector_name, "גבר",  pick_item(m_items, sector_idx), "החלפה יומית לפי צבע"])
    write_block("אמריקה הדרומית", SOUTH_AMERICA)
    write_block("אסיה", ASIA)

    # Style data rows
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
        for c in row:
            c.alignment = align
            c.border = border

    # Column widths
    widths = [22, 30, 38, 8, 36, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    wb.save(path)
    print(f"wrote {path}  ({ws.max_row-1} data rows)")

# ---------- Build stores sheet -----------------------------------------

def collect_all_items():
    """Walk the same data and collect (country, item) pairs preserving order, unique per pair."""
    seen = set()
    pairs = []  # (country_display, item)

    def add(country_display, item):
        key = (country_display, item)
        if key in seen:
            return
        seen.add(key)
        pairs.append((country_display, item))

    def walk(mapping, country_prefix=""):
        for country, sectors in mapping.items():
            display = f"{country_prefix}{country}" if country_prefix else country
            for sector_idx, (sector_name, culture) in enumerate(sectors):
                f_items, m_items = items_for(culture)
                add(display, pick_item(f_items, sector_idx))
                add(display, pick_item(m_items, sector_idx))

    walk(EUROPE)
    walk(NORTH_AMERICA_COUNTRIES)
    for state in US_STATES:
        sectors = us_state_sectors(state)
        for sector_idx, (sector_name, culture) in enumerate(sectors):
            f_items, m_items = items_for(culture)
            add(f"ארה\"ב - {state}", pick_item(f_items, sector_idx))
            add(f"ארה\"ב - {state}", pick_item(m_items, sector_idx))
    walk(SOUTH_AMERICA)
    walk(ASIA)
    return pairs

def build_stores_workbook(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "פריטים-חנויות"
    ws.sheet_view.rightToLeft = True

    headers = ["מדינה / מדינת ארה\"ב", "פריט", "שם חנות / רשת", "אתר רשמי"]
    ws.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="548235")
    align = Alignment(horizontal="right", vertical="center", wrap_text=True)
    border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    for col_idx, _ in enumerate(headers, 1):
        c = ws.cell(row=1, column=col_idx)
        c.font = header_font
        c.fill = header_fill
        c.alignment = align
        c.border = border

    pairs = collect_all_items()
    for country_display, item in pairs:
        # resolve store list
        if country_display.startswith("ארה\"ב - "):
            state = country_display.split(" - ", 1)[1]
            chains = us_stores_for(state)
        else:
            chains = stores_for(country_display)
        if not chains:
            ws.append([country_display, item, "(אין נתון מוודא)", ""])
            continue
        for name, url in chains:
            ws.append([country_display, item, name, url])

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
        for c in row:
            c.alignment = align
            c.border = border

    widths = [30, 32, 40, 50]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    wb.save(path)
    print(f"wrote {path}  ({ws.max_row-1} data rows)")

if __name__ == "__main__":
    build_main_workbook("countries_sectors_items.xlsx")
    build_stores_workbook("items_stores.xlsx")
